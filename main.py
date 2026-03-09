from fastapi import FastAPI, Response
from pydantic import BaseModel
import pandas as pd
import io
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

app = FastAPI()

class ReportData(BaseModel):
    operator_name: str
    assets: list[dict] 

@app.post("/report/xlsx")
async def generate_excel_report(data: ReportData):
    df = pd.DataFrame(data.assets)
    
    # 1. Have Pandas write the raw data to a temporary memory buffer
    raw_buffer = io.BytesIO()
    df.to_excel(raw_buffer, index=False, engine='openpyxl')
    raw_buffer.seek(0)
    
    # 2. Open that buffer strictly with openpyxl to apply bulletproof formatting
    wb = openpyxl.load_workbook(raw_buffer)
    ws = wb.active
    
    # Define the visual styles
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'), 
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    bold_font = Font(bold=True)
    
    # Red urgency styles (Light red background, dark red text)
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    red_font = Font(color='9C0006', bold=True)
    
    # Find the Urgency column index dynamically
    urgency_col_idx = None
    for col_idx, cell in enumerate(ws[1], start=1): # Row 1 is headers
        if cell.value == "Urgency":
            urgency_col_idx = col_idx
            break
            
    # Loop through every cell to apply borders, centering, and colors
    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        is_critical = False
        
        # Check if this row's urgency is critical
        if row_idx > 1 and urgency_col_idx:
            val = ws.cell(row=row_idx, column=urgency_col_idx).value
            if str(val).strip().lower() == "critical":
                is_critical = True
                
        for cell in row:
            # Apply table borders and center alignment to EVERY cell
            cell.border = thin_border
            cell.alignment = center_alignment
            
            if row_idx == 1:
                cell.font = bold_font  # Bold the headers
            elif is_critical:
                cell.fill = red_fill   # Highlight critical rows
                cell.font = red_font
                
    # Auto-size columns to perfectly fit the text
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 3
        
    # 3. Save the perfectly styled workbook to a new buffer and send it
    final_buffer = io.BytesIO()
    wb.save(final_buffer)
    final_buffer.seek(0)
    
    headers = {
        'Content-Disposition': f'attachment; filename="compliance_report_{data.operator_name}.xlsx"'
    }
    return Response(
        content=final_buffer.getvalue(), 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        headers=headers
    )
